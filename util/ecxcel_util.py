# -*- encoding: utf-8 -*-
'''
@File    : ecxcel_util.py
@Author  : vgzx
@Date    : 2020/08/14 22:28
@Contact : viewgrand@163.com
@Desc    : excel文件操作

'''

import openpyxl


e_path = "D:/project/stax_blog/doc/dev_log.xlsx"


def open():
    wb = openpyxl.load_workbook('./doc/dev_log.xlsx')
    print(wb.sheetnames)    # 取得全部的sheet名称
    sheet = wb['Sheet1']    # 根据sheet名称取得sheet对象
    # print(sheet.title)    # 取得页签名称

    # sheet.cell(row=2, column=1, coordinate = 'A2')   # cell拥有这三个属性，行列都是整型
    # sheet["A1"]    # 列从A开始，行从1开始  ["列行"]
    print(sheet['A2'].value)   # 取得指定单元对象，以及它的值

def write():
    wb2 = openpyxl.Workbook()      # 新建一个空的workbook对象
    ws2 = wb2.create_sheet(title = "数字", index = 0)   # 新建sheet  指定名称、位置
    for row in [(1, "lzx"), (2, 'hpx', 27), (3, 'cjw', '60kg')]:
        ws2.append(row)        # 可以插入列表、元祖、字典等对象
    wb2.save(filename="./doc/example.xlsx")

# open()

write()


